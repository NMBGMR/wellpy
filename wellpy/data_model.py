# ===============================================================================
# Copyright 2016 ross
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
# http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
# ===============================================================================
import os
import csv
import time

from datetime import datetime
from numpy import empty, polyfit, polyval, array, where, diff, logical_and, hstack, delete
from openpyxl import load_workbook

from wellpy.sigproc import smooth

fmts = ('%Y/%m/%d %H:%M:%S',
        '%Y/%m/%d %H:%M',
        '%m/%d/%Y %H:%M:%S',
        '%m/%d/%Y %H:%M',
        '%m/%d/%y %H:%M:%S',
        '%m/%d/%y %H:%M',
        '%Y-%m-%d %H:%M:%S',
        )


def extract_timestamp(date):

    for fmt in fmts:
        try:
            return datetime.strptime(date, fmt)
        except ValueError, e:
            continue

    else:
        raise ValueError('Invalid date format "{}"'
                         'Did not match any of the following\n{}'.format(date, '\n'.join(fmts)))


class Channel:
    identification = None
    reference_level = None
    value_range = None


class DataModel:
    adjusted_water_head = None
    water_head = None
    filtered_zeros = None
    serial_number = None
    pointid = None
    manual_water_depth_x = None
    manual_water_depth_y = None
    depth_to_water_x = None
    depth_to_water_y = None
    water_temp = None
    cond = None

    is_acoustic = False

    def __init__(self, path):
        self._path = path
        if os.path.isfile(path):
            self._load(path)

        self.manual_water_depth_x = array([])
        self.manual_water_depth_y = array([])
        self.water_depth_status = array([])
        self.depth_to_water_x = array([])
        self.depth_to_water_y = array([])
        self.omissions = []

    def get_owater_head(self):
        return array(self._owater_head)

    def get_water_head(self):
        return array(self._water_head)

    def set_water_head(self, v):
        self._water_head = v

    def get_depth_to_water(self):
        return array(self.depth_to_water_y)

    def smooth(self, ys, window, method, selection=None):
        
        if selection:
            rys = ys[:]
            x = self.x
            idx = where(logical_and(selection[0]<=x, x<=selection[1]))[0]
            
            pys = smooth(ys, window, method)
     
            rys[idx] = pys[idx]
        else:
            rys = smooth(ys, window, method)
            
        return rys
        
    def remove_up_spikes(self, ys, threshold, selection, normal_mode=True):
        x = self.x
        ys = array(ys[:])
        if selection:

            if normal_mode:
                m=None
                for i, xi in enumerate(x):
                    if selection[0]<=xi<=selection[1]:
                        if m is None:
                            m = ys[i]
                            continue
                        
                        if (m-ys[i])>threshold:
                            ys[i] = m
                        else:
                            m = ys[i]
            else:
                fidx = where(logical_and(selection[0]<=x, x<=selection[1]))[0]
                sidx = fidx[0]
                fys = ys[fidx]
                sys = smooth(fys, 11, 'hanning')    
                idx = sys - fys  >= threshold      
                #print('fidx', fidx)
                #print('sidx', sidx)
                #print('fys', fys)
                #print('max', max(fys))
                #print('idx', idx) 
                
                nidx = fidx[idx]
                ys[nidx] = sys[idx]
                
                #ys = max(fys)
            
                    
                    
            #fidx = where(logical_and(selection[0]<=x, x<=selection[1]))[0]
            #fys = ys[fidx]
            #idxs = where(diff(ys)>=threshold)[0]
            #print(idxs)
            #m = fys.max()
            #idxs = where(m-ys>threshold)[0]
            #for i in idxs:
            #    if selection[0] <= x[i] <= selection[1]:
            #        ys[i] = m
            
            #if idxs.any():
            #    if idxs.shape[0] == 1:
            #        idxs = array([idxs[0], ys.shape[0] - 1])

                
            #    for sidx in idxs:
            #        sx= x[sidx]
                    # print 'sxex', sx, ex
             #       if selection[0] <= sx <= selection[1]:
             #           ys
                #        ys[sidx] = ys[sidx-1]
        return ys
        
    def fix_data(self, ys, threshold, selection):
        x = self.x
        ys = array(ys[:])
        # find zeros
        zs = where(ys == 0)[0]
        fs = []

        # print 'selection', selection
        # if sx >= selection[0] and ex <= selection[1]:
        if selection:
            # mask = logical_and(x > selection[0], x < selection[1])
            # oys = ys[mask]
            # oxs = x[mask]

            idxs = where(abs(diff(ys)) >= threshold)[0]
            if idxs.any():
                if idxs.shape[0] == 1:
                    idxs = array([idxs[0], ys.shape[0] - 1])

                # n = idxs.shape[0]
                # if n % 2:
                #     idxs = hstack((idxs, [-1]))
                #     n += 1

                # for sidx, eidx in idxs.reshape(n / 2, 2):
                for sidx in idxs:
                    # sidx, eidx = idxs[0], idxs[1]
                    eidx = sidx + 1
                    sx, ex = x[sidx], x[eidx]
                    # print 'sxex', sx, ex
                    if sx >= selection[0] and ex <= selection[1]:
                        offset = ys[sidx] - ys[eidx]
                        fs.append((offset, sidx, eidx, sx, ex))
                        ys[sidx + 1:] += offset
        else:
            while 1:
                idxs = where(abs(diff(ys)) >= threshold)[0]
                if not idxs.any():
                    break
                elif idxs.shape[0] == 1:
                    idxs = [idxs[0], ys.shape[0] - 1]

                offset = ys[idxs[0]] - ys[idxs[0] + 1]

                sidx, eidx = idxs[0], idxs[1]
                sx, ex = x[sidx], x[eidx]
                fs.append((offset, sidx, eidx, sx, ex))
                ys[idxs[0] + 1:idxs[1] + 1] += offset
        return ys, zs, fs

    # private
    def _load(self, p):
        pp = p.lower()
        if pp.endswith('.csv'):
            self._load_csv(p)
        elif pp.endswith('.wcsv'):
            self._load_wcsv(p)
            self.is_acoustic = True
        else:
            self._load_xls(p)

    def _load_wcsv(self, p):
        #delimiter = ','
        # 2020-09-29 13:13:00
        x, y, ts = [], [], []
        self.pointid = os.path.splitext(os.path.basename(p))[0]
        
        with open(p, 'r') as rfile:
            dialect = csv.Sniffer().sniff(rfile.read(1024))
            rfile.seek(0)
            reader = csv.reader(rfile, dialect)
            for i, line in enumerate(reader):
            #for i, line in enumerate(rfile):
                oline = line
                #line = line.strip()
                #line = line.split(delimiter)
                if len(line) == 4:
                    date, tempC, tempR, depth = line
                    try:
                        depth = float(depth)
                    except (ValueError, TypeError):
                        continue

                    try:
                        temp = float(tempC)
                    except (ValueError, TypeError):
                        continue

                    date = extract_timestamp(date)

                    x.append(time.mktime(date.timetuple()))
                    y.append(depth)
                    # ws.append(water_head)
                    ts.append(temp)
                    # cs.append(cond)
        self.x = array(x)
        self.depth_to_water_x = x
        self.raw_depth_to_water_y = array(y)
        self.depth_to_water_y = array(y)
        self.temp_air = ts

    def _load_csv(self, p):
        delimiter = ','
        x = []
        ws = []
        ts = []
        cs = []

        cfmt = None
        with open(p, 'r') as rfile:
            for i, line in enumerate(rfile):
                oline = line
                line = line.strip()
                line = line.split(delimiter)

                print 'line i={}, len={}, {}'.format(i, len(line), line)

                if not self.serial_number:
                    if line[0].startswith('Serial number'):
                        self.serial_number = line[0].split('=')[1].strip()

                if not self.pointid:
                    if line[0].startswith('Location'):
                        self.pointid = line[0].split('=')[1].strip().upper()

                if i < 53:
                    continue

                cond = 0
                if len(line) == 3:
                    date, water_head, temp = line
                elif len(line) == 4:
                    date, water_head, temp, cond = line
                else:
                    if oline.startswith('END OF DATA'):
                        break

                try:
                    water_head = float(water_head)
                except (ValueError, TypeError):
                    continue
                try:
                    temp = float(temp)
                except (ValueError, TypeError):
                    continue
                try:
                    cond = float(cond)
                except (ValueError, TypeError):
                    continue

                if not cfmt:
                   date = extract_timestamp(date)
                else:
                    date = datetime.strptime(date, cfmt)

                x.append(time.mktime(date.timetuple()))
                ws.append(water_head)
                ts.append(temp)
                cs.append(cond)

        self.x = array(x)

        # ws = array(ws)
        self.water_temp = array(ts)
        self._water_head = ws
        self._owater_head = ws
        self.water_head = array(ws)
        self.adjusted_water_head = array(ws)
        self.cond = array(cs)

    def _load_xls(self, p):
        wb = load_workbook(p, data_only=True)
        sheet = wb[wb.sheetnames[0]]

        # set_location(data, sheet)
        def extract_cell_value(idx):
            rv = sheet[idx].value
            v = rv.split('=')[1]
            return v

        for attr, idx, func in (('location', 'A16', None),
                                ('sample_period', 'A17', None),
                                ('sample_method', 'A18', None)):
            v = extract_cell_value(idx)
            if func:
                v = func(v)
            setattr(self, attr, v)

        # read in channels
        # right now just hard code in two channels

        ch1 = Channel()
        ch1.identification = extract_cell_value('A21')
        ch1.reference_level = extract_cell_value('A22')
        ch1.value_range = extract_cell_value('A23')

        ch2 = Channel()
        ch2.identification = extract_cell_value('A27')
        ch2.reference_level = extract_cell_value('A28')
        ch2.value_range = extract_cell_value('A29')

        self.channels = (ch1, ch2)

        # load the time series
        n = sheet.max_row - 52
        x = empty(n)
        t = empty(n)
        wh = empty(n)
        awh = empty(n)
        wle = empty(n)

        arrs = (x, t, wh, awh, wle)
        for i, row in enumerate(sheet.iter_rows(min_row=53)):
            for j, arr in enumerate(arrs):
                v = row[j].value
                if j == 0:
                    v = time.mktime(v.timetuple())
                arr[i] = v
            if i > 10:
                break

        self.x = x
        self.temp = t
        self.water_head = wh
        self.adjusted_water_head = awh
        self.water_level_elevation = wle

        # ============= EOF =============================================
        # def apply_linear(self, attr, s, e, mask):
        #     v = getattr(self, attr)
        #     ys = v[mask]
        #     xs = self.x[mask]
        #
        #     fxs = (xs[0], xs[-1])
        #     fys = (ys[0], ys[-1])
        #     coeffs = polyfit(fxs, fys, 1)
        #     v[mask] = polyval(coeffs, xs)
        #
        #
        # def apply_offset(self, attr, offset, mask):
        #     v = getattr(self, attr)
        #     v[mask] += offset
